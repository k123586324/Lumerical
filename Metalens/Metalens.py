from openpyxl import load_workbook
import numpy as np
from matplotlib.pylab import mpl
import math
import matplotlib.pyplot as plt

def search_pillar_r():
    # 讀取 Excel 檔案
    wb = load_workbook(r'C:\Users\AlfieLiu\Desktop\Metalens_test_2.xlsx')
    sheet = wb['工作表3']
    # 根據位置取得儲存格
    # 讀取角度
    columns = sheet.columns
    deg = []  # 用於存放一行資料
    for i in list(columns)[1]:  # 遍歷每行資料
        deg.append(i.value)
    # 讀取直徑
    columns = sheet.columns
    pillar_r = []
    for i in list(columns)[0]:
        pillar_r.append(i.value)
        # print(i.value)

    # 半徑(mm)
    semi_diameter =0.015
    # 取樣數
    sampling = 60
    # Zemax Norm Radius
    norm_radius = 2
    # 係數
    coefficients_Zemax = [-3.930182550882000E+005, 1.184068070989000E+008, -7.101677164762000E+010,
                          4.739775438444000E+013]
    # 係數轉換
    coefficients = [0] * len(coefficients_Zemax)
    for i in range(len(coefficients_Zemax)):
        coefficients[i] = coefficients_Zemax[i] / (2 * math.pi * norm_radius ** ((i + 1) * 2))

    phase_1D = []
    r_list = []
    for i in range(sampling + 1):
        r = semi_diameter * (i) / sampling
        phase = 0
        for c in range(len(coefficients) - 1, -1, -1):
            phase = (phase + coefficients[c]) * r ** 2
        #print(f'{i},{r},{phase}')
        phase_1D.append(phase)
        r_list.append(r)

    # plt.plot(r_list, phase_1D)
    # plt.show()

    for i in phase_1D:
        phase_deg = (i % 1) * 360
        min = 9999
        index=0
        index_min=0
        for j in deg:
            delta = abs(phase_deg - j)
            if delta < min:
                min=delta
                index_min=index
            index+=1
        print(pillar_r[index_min])

def Zemax_grid_phase():
    # 讀取exccel和txt
    wb = load_workbook('output.xlsx')
    sheet = wb['工作表1']
    path = 'output.txt'
    f = open(path, 'w')
    # 半徑(mm)
    semi_diameter = 0.6762988
    # 取樣數
    sampling = 64
    pitch=(semi_diameter*2)/(sampling*2+1)
    # 趨勢線方程式
    coefficients_Zemax = [-4.513024,
                          11.453339,
                          -474.037022,
                          0.765004,
                          -0.017370]
    # for i in range(sampling + 1):
    #     x=semi_diameter * (i) / sampling
    #     y=coefficients_Zemax[0]*x**4+coefficients_Zemax[1]*x**3+coefficients_Zemax[2]*x**2+coefficients_Zemax[3]*x+coefficients_Zemax[4]
    #     print(y)
    f.write(f'{(sampling*2+1)} {(sampling*2+1)} {pitch} {pitch} 0\n')
    for i in range(sampling*2 + 1):
        for j in range(sampling * 2 + 1):
            k=((i-sampling)**2+(j-sampling)**2)**0.5
            x = semi_diameter * (k) / sampling
            if x<=semi_diameter:
                y = coefficients_Zemax[0] * x ** 4 + coefficients_Zemax[1] * x ** 3 + coefficients_Zemax[2] * x ** 2 + \
                    coefficients_Zemax[3] * x + coefficients_Zemax[4]
                sheet.cell(row=i + 1, column=j + 1, value=(y*2*np.pi))
                f.write(f'{y * 2 * np.pi}\n')
                print(y * 2 * np.pi)
                #print(f'{i},{j},{k},{x}')
            else:
                sheet.cell(row=i + 1, column=j + 1, value=0)
                f.write("0\n")
                print(0)
    wb.save('output.xlsx')
    f.close()

def test2():
    sampling=65
    wb = load_workbook('output.xlsx')
    sheet = wb['工作表1']
    f = open('test.txt')
    k = f.readlines()
    #print(k[0].split(' ')[0])
    i=1
    j=1
    for phase in k:
        print(phase.split('\n')[0])
        sheet.cell(row=i, column=j, value=(float(phase.split('\n')[0])))
        i+=1
        if i>sampling:
            i=1
            j+=1
    f.close()
    wb.save('output.xlsx')

def test3():
    sampling=65
    wb = load_workbook('s4_15deg.xlsx')
    sheet = wb['工作表1']
    fig = plt.figure(figsize=(10, 10))
    #ax3d = plt.axes(projection="3d")
    xdata = np.linspace(-0.578928336, 0.578928336, sampling)
    ydata = np.linspace(-0.625304966, 0.524209102, sampling)
    X, Y = np.meshgrid(xdata, ydata)
    Z = np.zeros((sampling, sampling))
    for i in range(sampling):
        for j in range(sampling):
            c = sheet.cell(row=(j+1), column=(i+1))
            Z[i][j] = c.value
    ax3d = plt.axes(projection='3d')
    surf=ax3d.plot_surface(X, Y, Z, cmap='jet')
    fig.colorbar(surf, ax=ax3d)
    ax3d.set_title('Surface Plot in Matplotlib')
    ax3d.set_xlabel('X')
    ax3d.set_ylabel('Y')
    ax3d.set_zlabel('Z')
    plt.show()


#test2()
test3()
#Zemax_grid_phase()

